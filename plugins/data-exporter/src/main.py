"""
数据导出器插件
支持将数据导出为CSV、JSON、Excel等格式
"""
import json
import csv
import io
from typing import List, Dict, Any, Optional

# 插件配置
_config = {
    "max_rows": 10000,
    "default_format": "csv"
}

# 支持的格式
SUPPORTED_FORMATS = {
    "csv": {"name": "CSV", "mime": "text/csv", "extension": ".csv"},
    "json": {"name": "JSON", "mime": "application/json", "extension": ".json"},
    "excel": {"name": "Excel", "mime": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "extension": ".xlsx"},
    "markdown": {"name": "Markdown", "mime": "text/markdown", "extension": ".md"}
}

def on_activate(context: dict):
    """插件激活时调用"""
    global _config
    if "config" in context:
        _config.update(context["config"])
    print(f"[数据导出器] 已激活，配置: {_config}")

def on_deactivate():
    """插件停用时调用"""
    print("[数据导出器] 已停用")

def register_routes(app):
    """注册API路由"""
    from fastapi import HTTPException
    from fastapi.responses import StreamingResponse
    from pydantic import BaseModel
    from typing import List, Dict, Any
    
    class ExportInput(BaseModel):
        data: List[Dict[str, Any]]
        format: Optional[str] = None
        filename: Optional[str] = None
    
    @app.get("/api/plugins/data-exporter/formats")
    async def list_formats():
        """获取支持的导出格式"""
        return {
            "status": "ok",
            "data": {
                "formats": SUPPORTED_FORMATS,
                "default": _config["default_format"]
            }
        }
    
    @app.post("/api/plugins/data-exporter/export")
    async def export_data(input_data: ExportInput):
        """导出数据"""
        if not input_data.data:
            raise HTTPException(400, "数据不能为空")
        
        if len(input_data.data) > _config["max_rows"]:
            raise HTTPException(400, f"数据超过最大行数限制({_config['max_rows']})")
        
        fmt = input_data.format or _config["default_format"]
        if fmt not in SUPPORTED_FORMATS:
            raise HTTPException(400, f"不支持的格式: {fmt}，支持: {list(SUPPORTED_FORMATS.keys())}")
        
        filename = input_data.filename or f"export{SUPPORTED_FORMATS[fmt]['extension']}"
        
        if fmt == "csv":
            content = _export_csv(input_data.data)
            return StreamingResponse(
                io.BytesIO(content.encode('utf-8-sig')),
                media_type="text/csv",
                headers={"Content-Disposition": f"attachment; filename={filename}"}
            )
        elif fmt == "json":
            content = _export_json(input_data.data)
            return StreamingResponse(
                io.BytesIO(content.encode('utf-8')),
                media_type="application/json",
                headers={"Content-Disposition": f"attachment; filename={filename}"}
            )
        elif fmt == "markdown":
            content = _export_markdown(input_data.data)
            return StreamingResponse(
                io.BytesIO(content.encode('utf-8')),
                media_type="text/markdown",
                headers={"Content-Disposition": f"attachment; filename={filename}"}
            )
        elif fmt == "excel":
            content = _export_excel(input_data.data)
            return StreamingResponse(
                io.BytesIO(content),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename={filename}"}
            )
    
    print("[数据导出器] 路由已注册: GET /api/plugins/data-exporter/formats, POST /export")

def _export_csv(data: List[Dict[str, Any]]) -> str:
    """导出为CSV"""
    if not data:
        return ""
    
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=data[0].keys())
    writer.writeheader()
    writer.writerows(data)
    return output.getvalue()

def _export_json(data: List[Dict[str, Any]]) -> str:
    """导出为JSON"""
    return json.dumps(data, ensure_ascii=False, indent=2)

def _export_markdown(data: List[Dict[str, Any]]) -> str:
    """导出为Markdown表格"""
    if not data:
        return ""
    
    headers = data[0].keys()
    lines = []
    
    # 表头
    lines.append("| " + " | ".join(headers) + " |")
    lines.append("| " + " | ".join(["---"] * len(headers)) + " |")
    
    # 数据行
    for row in data:
        lines.append("| " + " | ".join(str(row.get(h, "")) for h in headers) + " |")
    
    return "\n".join(lines)

def _export_excel(data: List[Dict[str, Any]]) -> bytes:
    """导出为Excel"""
    try:
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        
        if not data:
            return b""
        
        # 写入表头
        headers = list(data[0].keys())
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # 写入数据
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, header in enumerate(headers, 1):
                ws.cell(row=row_idx, column=col_idx, value=row_data.get(header))
        
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()
    except ImportError:
        # 如果没有openpyxl，返回JSON格式
        return _export_json(data).encode('utf-8')

def health_check():
    """健康检查"""
    return {"status": "ok", "plugin": "data-exporter", "version": "1.0.0"}
