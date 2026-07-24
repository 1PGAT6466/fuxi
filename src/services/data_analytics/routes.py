"""
routes.py — 数据分析服务 API 路由（v2.2 增强版）

端点:
  POST /api/analytics/stats        — 综合统计
  POST /api/analytics/trends       — 趋势数据
  POST /api/analytics/report       — 统计报表
  POST /api/analytics/storage      — 存储分布
  POST /api/analytics/export       — 多格式导出（PDF/Excel/CSV/JSON）
  GET  /api/analytics/health       — 健康检查
  GET  /api/analytics/templates    — 报表模板列表
  POST /api/analytics/templates    — 创建模板
  PUT  /api/analytics/templates/{id}  — 更新模板
  DELETE /api/analytics/templates/{id} — 删除模板
  POST /api/reports/share          — 生成分享链接
  GET  /api/reports/{token}        — 访问分享报表
  DELETE /api/reports/share/{token} — 撤销分享
"""

import csv
import io
import json
import logging
import sqlite3
import uuid
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional, List

from fastapi import APIRouter, HTTPException, Query, Body
from pydantic import BaseModel, Field

# ── 从项目配置获取路径 ──
from src.config import DATA_DIR, LOG_DIR

logger = logging.getLogger("services.data-analytics.routes")

# ── 分享数据存储路径 ──
SHARED_REPORTS_DIR = DATA_DIR / "shared_reports"
TEMPLATES_DIR = DATA_DIR / "export_templates"

router = APIRouter(prefix="", tags=["Data Analytics"])

# ── Pydantic 请求模型 ──

class TrendsRequest(BaseModel):
    """趋势查询请求"""
    period: str = Field(
        default="day",
        description="时间维度: day / week / month",
        examples=["day", "week", "month"]
    )
    start_date: Optional[str] = Field(
        default=None,
        description="起始日期 (YYYY-MM-DD)，不填则默认最近 30 天",
        examples=["2026-01-01"]
    )
    end_date: Optional[str] = Field(
        default=None,
        description="结束日期 (YYYY-MM-DD)，不填则默认今天",
        examples=["2026-01-31"]
    )


class ReportRequest(BaseModel):
    """报表请求"""
    report_type: str = Field(
        default="summary",
        description="报表类型: summary / detailed",
        examples=["summary", "detailed"]
    )
    period: str = Field(
        default="month",
        description="时间范围: today / week / month / all",
        examples=["month"]
    )


class ExportRequest(BaseModel):
    """导出请求（v2.2 增强版：支持 PDF/Excel/CSV/JSON）"""
    format: str = Field(
        default="csv",
        description="导出格式: pdf / excel / csv / json",
        examples=["csv", "excel", "pdf", "json"]
    )
    fields: List[str] = Field(
        default=[],
        description="导出字段列表",
        examples=[["date", "queries", "documents", "users"]]
    )
    date_range: Optional[dict] = Field(
        default=None,
        description="时间范围 {start, end}",
        examples=[{"start": "2026-01-01", "end": "2026-06-30"}]
    )
    template_id: Optional[str] = Field(
        default=None,
        description="报表模板 ID（可选）"
    )
    title: Optional[str] = Field(
        default=None,
        description="导出标题（可选）"
    )


class TemplateCreateRequest(BaseModel):
    """创建模板请求"""
    name: str = Field(description="模板名称")
    description: str = Field(default="", description="模板描述")
    default_fields: List[str] = Field(description="预设字段列表")
    default_format: str = Field(default="csv", description="预设导出格式")


class TemplateUpdateRequest(BaseModel):
    """更新模板请求"""
    name: Optional[str] = None
    description: Optional[str] = None
    default_fields: Optional[List[str]] = None
    default_format: Optional[str] = None


class ShareRequest(BaseModel):
    """报表分享请求"""
    report_id: str = Field(description="报表 ID")
    permissions: List[str] = Field(
        default=["view"],
        description="权限列表: view / edit / download"
    )
    expires_at: Optional[str] = Field(
        default=None,
        description="过期时间（ISO 格式）"
    )
    password: Optional[str] = Field(
        default=None,
        description="访问密码（6-20 位）"
    )
    note: Optional[str] = Field(
        default=None,
        description="分享备注"
    )


# ══════════════════════════════════════════════════
#  数据采集层 — 从各数据源提取统计信息
# ══════════════════════════════════════════════════

def _get_chroma_stats() -> dict:
    """从 ChromaDB 获取向量存储统计"""
    chroma_path = DATA_DIR / "chromadb" / "chroma.sqlite3"
    result = {
        "collections": 0,
        "embeddings": 0,
        "vectors_total": 0,
        "storage_size_bytes": 0,
    }
    try:
        if chroma_path.exists():
            conn = sqlite3.connect(str(chroma_path))
            conn.row_factory = sqlite3.Row
            cur = conn.cursor()

            # 集合数
            cur.execute("SELECT COUNT(*) as cnt FROM collections")
            row = cur.fetchone()
            result["collections"] = row["cnt"] if row else 0

            # 嵌入向量数 (从 FTS data 估算)
            cur.execute("SELECT COUNT(*) as cnt FROM embedding_fulltext_search_data")
            row = cur.fetchone()
            fts_count = row["cnt"] if row else 0
            result["embeddings"] = fts_count

            # 向量总数（segment 级别的 counting）
            cur.execute("SELECT COUNT(*) as cnt FROM segments WHERE type='VECTOR'")
            row = cur.fetchone()
            seg_count = row["cnt"] if row else 0
            result["vectors_total"] = fts_count

            conn.close()

        # 存储大小：chroma 目录总大小
        chroma_dir = DATA_DIR / "chromadb"
        if chroma_dir.exists():
            total_bytes = sum(
                f.stat().st_size
                for f in chroma_dir.rglob("*")
                if f.is_file()
            )
            result["storage_size_bytes"] = total_bytes

    except Exception as e:  # TODO: Narrow exception type
        logger.warning(f"获取 ChromaDB 统计失败: {e}")

    return result


def _get_chunks_stats() -> dict:
    """从 chunks.db 获取文档分块统计"""
    from src.config import CHUNKS_DB_PATH

    result = {
        "total_chunks": 0,
        "total_documents": 0,
        "active_chunks": 0,
        "categories": {},
    }
    try:
        db_path = Path(CHUNKS_DB_PATH)
        if db_path.exists():
            conn = sqlite3.connect(str(db_path))
            conn.row_factory = sqlite3.Row
            cur = conn.cursor()

            # 总数
            cur.execute("SELECT COUNT(*) as cnt FROM chunks")
            row = cur.fetchone()
            result["total_chunks"] = row["cnt"] if row else 0

            # 活跃
            cur.execute("SELECT COUNT(*) as cnt FROM chunks WHERE status='active'")
            row = cur.fetchone()
            result["active_chunks"] = row["cnt"] if row else 0

            # 按 category 分布
            cur.execute(
                "SELECT category, COUNT(*) as cnt FROM chunks "
                "GROUP BY category ORDER BY cnt DESC"
            )
            cats = {}
            for row in cur.fetchall():
                cat = row["category"] or "未分类"
                cats[cat] = row["cnt"]
            result["categories"] = cats

            # 不同文档数（按 file_hash 去重）
            cur.execute(
                "SELECT COUNT(DISTINCT file_hash) as cnt FROM chunks WHERE file_hash IS NOT NULL"
            )
            row = cur.fetchone()
            result["total_documents"] = row["cnt"] if row else 0

            conn.close()

    except Exception as e:  # TODO: Narrow exception type
        logger.warning(f"获取 chunks 统计失败: {e}")

    return result


def _get_users_stats() -> dict:
    """从用户数据获取统计"""
    result = {"total_users": 0, "active_users": 0, "roles": {}}
    try:
        users_file = DATA_DIR / "users.json"
        if users_file.exists():
            data = json.loads(users_file.read_text(encoding="utf-8"))
            if isinstance(data, dict):
                result["total_users"] = len(data)
                if "roles" in data:
                    result["roles"] = data["roles"]
            elif isinstance(data, list):
                result["total_users"] = len(data)
    except Exception as e:  # TODO: Narrow exception type
        logger.warning(f"获取用户统计失败: {e}")

    return result


def _get_search_logs_stats(start_date: Optional[str] = None, end_date: Optional[str] = None) -> list:
    """从搜索日志中提取查询记录（按天聚合）"""
    records = []
    try:
        # 读取搜索日志 JSONL 文件
        log_dir = Path(LOG_DIR)
        if not log_dir.exists():
            return records

        # 如果没有指定日期范围，默认最近 30 天
        if not end_date:
            end_dt = datetime.now()
        else:
            end_dt = datetime.strptime(end_date, "%Y-%m-%d")
        if not start_date:
            start_dt = end_dt - timedelta(days=30)
        else:
            start_dt = datetime.strptime(start_date, "%Y-%m-%d")

        # 按天聚合
        daily_counts = {}
        current = start_dt
        while current <= end_dt:
            key = current.strftime("%Y-%m-%d")
            daily_counts[key] = {"count": 0, "total_latency": 0.0}
            current += timedelta(days=1)

        # 读取日志文件
        log_files = list(log_dir.glob("search_*.jsonl"))
        for log_file in log_files:
            try:
                content = log_file.read_text(encoding="utf-8")
                for line in content.strip().split("\n"):
                    if not line.strip():
                        continue
                    try:
                        entry = json.loads(line)
                        ts = entry.get("timestamp", "")
                        date_key = ts[:10] if ts else None
                        if date_key and date_key in daily_counts:
                            daily_counts[date_key]["count"] += 1
                            daily_counts[date_key]["total_latency"] += entry.get(
                                "latency_ms", 0
                            )
                    except json.JSONDecodeError:
                        continue
            except Exception as e:  # TODO: Narrow exception type
                logger.warning("解析搜索日志条目失败: %s", e, exc_info=True)
                continue

        records = [
            {
                "date": k,
                "query_count": v["count"],
                "avg_latency_ms": round(v["total_latency"] / v["count"], 1)
                if v["count"] > 0
                else 0,
            }
            for k, v in sorted(daily_counts.items())
        ]

    except Exception as e:  # TODO: Narrow exception type
        logger.warning(f"获取搜索日志统计失败: {e}")

    return records


def _get_uploads_stats() -> dict:
    """从 uploads 目录获取文件统计"""
    result = {
        "total_files": 0,
        "total_size_bytes": 0,
        "by_extension": {},
    }
    try:
        upload_dir = DATA_DIR / "uploads"
        if upload_dir.exists():
            extensions = {}
            total_bytes = 0
            total_files = 0
            for f in upload_dir.rglob("*"):
                if f.is_file():
                    total_files += 1
                    size = f.stat().st_size
                    total_bytes += size
                    ext = f.suffix.lower() or "无后缀"
                    if ext not in extensions:
                        extensions[ext] = {"count": 0, "size_bytes": 0}
                    extensions[ext]["count"] += 1
                    extensions[ext]["size_bytes"] += size
            result["total_files"] = total_files
            result["total_size_bytes"] = total_bytes
            result["by_extension"] = extensions
    except Exception as e:  # TODO: Narrow exception type
        logger.warning(f"获取上传文件统计失败: {e}")

    return result


def _get_audit_stats() -> dict:
    """从 audit.db 获取审计/API 调用统计"""
    result = {
        "total_events": 0,
        "api_calls": {},
    }
    try:
        audit_path = DATA_DIR / "audit.db"
        if audit_path.exists():
            conn = sqlite3.connect(str(audit_path))
            conn.row_factory = sqlite3.Row
            cur = conn.cursor()

            # 检查表结构
            cur.execute("SELECT name FROM sqlite_master WHERE type='table'")
            tables = [r[0] for r in cur.fetchall()]
            logger.debug(f"audit.db tables: {tables}")

            if "audit_events" in tables:
                cur.execute("SELECT COUNT(*) as cnt FROM audit_events")
                row = cur.fetchone()
                result["total_events"] = row["cnt"] if row else 0

                # 按端点聚合
                cur.execute(
                    "SELECT endpoint, COUNT(*) as cnt FROM audit_events "
                    "GROUP BY endpoint ORDER BY cnt DESC"
                )
                api_calls = {}
                for row in cur.fetchall():
                    api_calls[row["endpoint"]] = row["cnt"]
                result["api_calls"] = api_calls

            conn.close()
    except Exception as e:  # TODO: Narrow exception type
        logger.warning(f"获取审计统计失败: {e}")

    return result


# ══════════════════════════════════════════════════
#  API 端点
# ══════════════════════════════════════════════════

# ── 健康检查 ──
@router.get("/health")
def health_check():
    """服务健康检查"""
    import platform
    from src.config import VERSION, START_TIME
    import time

    uptime_seconds = int(time.time() - START_TIME)

    return {
        "status": "ok",
        "service": "data-analytics",
        "version": "1.0.0",
        "platform_version": VERSION,
        "uptime_seconds": uptime_seconds,
        "python_version": platform.python_version(),
        "timestamp": datetime.now().isoformat(),
    }


# ── GET /api/analytics/stats 也支持（方便浏览器调试）──
@router.get("/stats")
def get_stats_get():
    """综合统计 (GET)"""
    return _build_stats()


# ── POST /api/analytics/stats — 综合统计 ──
@router.post("/stats")
def get_stats():
    """
    综合统计
    返回文档数、用户数、向量数、存储大小等综合信息
    """
    return _build_stats()


def _build_stats() -> dict:
    """构建综合统计数据"""
    chroma = _get_chroma_stats()
    chunks = _get_chunks_stats()
    users = _get_users_stats()
    uploads = _get_uploads_stats()
    audit = _get_audit_stats()

    return {
        "timestamp": datetime.now().isoformat(),
        "vectors": {
            "collections": chroma["collections"],
            "embeddings": chroma["embeddings"],
            "vectors_total": chroma["vectors_total"],
            "storage_size_bytes": chroma["storage_size_bytes"],
            "storage_size_mb": round(chroma["storage_size_bytes"] / (1024 * 1024), 2),
        },
        "documents": {
            "total_chunks": chunks["total_chunks"],
            "active_chunks": chunks["active_chunks"],
            "total_documents": chunks["total_documents"],
            "categories": chunks["categories"],
        },
        "users": {
            "total_users": users["total_users"],
            "active_users": users.get("active_users", 0),
            "roles": users.get("roles", {}),
        },
        "storage": {
            "total_files": uploads["total_files"],
            "total_size_bytes": uploads["total_size_bytes"],
            "total_size_mb": round(uploads["total_size_bytes"] / (1024 * 1024), 2),
        },
        "audit": {
            "total_events": audit["total_events"],
            "api_calls": audit.get("api_calls", {}),
        },
    }


# ── POST /api/analytics/trends — 趋势数据 ──
@router.post("/trends")
def get_trends(request: TrendsRequest):
    """
    趋势数据
    按日/周/月维度返回文档增长、查询量、API 调用量
    """
    period = request.period
    if period not in ("day", "week", "month"):
        raise HTTPException(400, "period 必须是 day / week / month 之一")

    search_logs = _get_search_logs_stats(request.start_date, request.end_date)

    # 按维度聚合
    if period == "month":
        aggregated = {}
        for entry in search_logs:
            month_key = entry["date"][:7]  # YYYY-MM
            if month_key not in aggregated:
                aggregated[month_key] = {
                    "period": month_key,
                    "query_count": 0,
                    "total_latency": 0.0,
                }
            aggregated[month_key]["query_count"] += entry["query_count"]
            aggregated[month_key]["total_latency"] += entry["avg_latency_ms"] * entry["query_count"]
        records = []
        for v in aggregated.values():
            records.append({
                "period": v["period"],
                "query_count": v["query_count"],
                "avg_latency_ms": round(
                    v["total_latency"] / v["query_count"], 1
                ) if v["query_count"] > 0 else 0,
            })
        records.sort(key=lambda x: x["period"])
    elif period == "week":
        aggregated = {}
        for entry in search_logs:
            dt = datetime.strptime(entry["date"], "%Y-%m-%d")
            iso_week = dt.isocalendar()
            week_key = f"{iso_week[0]}-W{iso_week[1]:02d}"
            if week_key not in aggregated:
                aggregated[week_key] = {
                    "period": week_key,
                    "query_count": 0,
                    "total_latency": 0.0,
                }
            aggregated[week_key]["query_count"] += entry["query_count"]
            aggregated[week_key]["total_latency"] += entry["avg_latency_ms"] * entry["query_count"]
        records = []
        for v in aggregated.values():
            records.append({
                "period": v["period"],
                "query_count": v["query_count"],
                "avg_latency_ms": round(
                    v["total_latency"] / v["query_count"], 1
                ) if v["query_count"] > 0 else 0,
            })
        records.sort(key=lambda x: x["period"])
    else:
        records = [
            {
                "period": e["date"],
                "query_count": e["query_count"],
                "avg_latency_ms": e["avg_latency_ms"],
            }
            for e in search_logs
        ]

    # 获取 chunk 创建趋势（按日聚合）
    chunk_trends = _get_chunk_creation_trends(request.start_date, request.end_date)

    return {
        "timestamp": datetime.now().isoformat(),
        "period": period,
        "queries": records,
        "documents": chunk_trends,
        "total_days": len(search_logs) if search_logs else 0,
    }


def _get_chunk_creation_trends(start_date: Optional[str] = None, end_date: Optional[str] = None) -> list:
    """从 chunks.db 提取文档创建趋势"""
    from src.config import CHUNKS_DB_PATH

    records = []
    try:
        db_path = Path(CHUNKS_DB_PATH)
        if not db_path.exists():
            return records

        if not end_date:
            end_dt = datetime.now()
        else:
            end_dt = datetime.strptime(end_date, "%Y-%m-%d")
        if not start_date:
            start_dt = end_dt - timedelta(days=30)
        else:
            start_dt = datetime.strptime(start_date, "%Y-%m-%d")

        # 初始化每天
        daily = {}
        current = start_dt
        while current <= end_dt:
            key = current.strftime("%Y-%m-%d")
            daily[key] = 0
            current += timedelta(days=1)

        conn = sqlite3.connect(str(db_path))
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        cur.execute(
            "SELECT DATE(created_at) as d, COUNT(*) as cnt FROM chunks "
            "WHERE created_at IS NOT NULL "
            "GROUP BY DATE(created_at) ORDER BY d"
        )
        for row in cur.fetchall():
            d = row["d"]
            if d in daily:
                daily[d] += row["cnt"]
        conn.close()

        records = [
            {"date": k, "new_documents": v}
            for k, v in sorted(daily.items())
        ]

    except Exception as e:  # TODO: Narrow exception type
        logger.warning(f"获取 chunk 趋势失败: {e}")

    return records


# ── POST /api/analytics/report — 统计报表 ──
@router.post("/report")
def generate_report(request: ReportRequest):
    """
    生成统计报表
    返回结构化数据供前端渲染
    """
    report_type = request.report_type
    if report_type not in ("summary", "detailed"):
        raise HTTPException(400, "report_type 必须是 summary / detailed")

    base_stats = _build_stats()
    report = {
        "generated_at": datetime.now().isoformat(),
        "report_type": report_type,
        "period": request.period,
        "summary": {
            "total_documents": base_stats["documents"]["total_documents"],
            "total_chunks": base_stats["documents"]["total_chunks"],
            "active_chunks": base_stats["documents"]["active_chunks"],
            "total_users": base_stats["users"]["total_users"],
            "vector_collections": base_stats["vectors"]["collections"],
            "vector_embeddings": base_stats["vectors"]["embeddings"],
            "storage_mb": base_stats["vectors"]["storage_size_mb"],
        },
    }

    if report_type == "detailed":
        report["details"] = {
            "document_categories": base_stats["documents"]["categories"],
            "storage_by_extension": {},
            "api_usage": base_stats.get("audit", {}).get("api_calls", {}),
            "queries_trend": _get_search_logs_stats(),
            "chunk_trend": _get_chunk_creation_trends(),
        }
        # 添加上传文件扩展名分布
        uploads = _get_uploads_stats()
        report["details"]["storage_by_extension"] = uploads["by_extension"]

    return report


# ── POST /api/analytics/storage — 存储分布 ──
@router.post("/storage")
def get_storage_distribution():
    """
    存储分布
    按文档类型、按用户等维度展示存储分布
    """
    chunks = _get_chunks_stats()
    uploads = _get_uploads_stats()

    # 按文档类型（后缀名）
    extensions = []
    for ext, info in sorted(
        uploads["by_extension"].items(),
        key=lambda x: x[1]["size_bytes"],
        reverse=True,
    ):
        extensions.append({
            "extension": ext,
            "file_count": info["count"],
            "size_bytes": info["size_bytes"],
            "size_mb": round(info["size_bytes"] / (1024 * 1024), 2),
        })

    # 按类别
    categories = [
        {"category": cat, "chunk_count": count}
        for cat, count in sorted(
            chunks["categories"].items(), key=lambda x: x[1], reverse=True
        )
    ]

    # 向量存储大小
    chroma = _get_chroma_stats()

    return {
        "timestamp": datetime.now().isoformat(),
        "by_extension": extensions,
        "by_category": categories,
        "vector_storage": {
            "size_bytes": chroma["storage_size_bytes"],
            "size_mb": round(chroma["storage_size_bytes"] / (1024 * 1024), 2),
            "collections": chroma["collections"],
            "embeddings": chroma["embeddings"],
        },
        "total_uploads_size_mb": round(uploads["total_size_bytes"] / (1024 * 1024), 2),
    }


# ── POST /api/analytics/export — 多格式导出（v2.2 增强版）──
@router.post("/export")
def export_report_v2(request: ExportRequest):
    """
    多格式数据导出：PDF / Excel / CSV / JSON
    支持字段筛选、时间范围过滤
    """
    fmt = request.format.lower()
    if fmt not in ("pdf", "excel", "csv", "json"):
        raise HTTPException(400, "format 必须是 pdf / excel / csv / json 之一")

    stats = _build_stats()
    now_str = datetime.now().strftime("%Y%m%d_%H%M%S")
    title = request.title or f"数据导出_{now_str}"

    # 收集导出数据
    export_data = _build_export_data(stats)

    # 字段过滤
    if request.fields:
        export_data = {k: v for k, v in export_data.items() if k in request.fields}

    if fmt == "json":
        return _export_json(export_data, title, now_str)
    elif fmt == "csv":
        return _export_csv(export_data, title, now_str)
    elif fmt == "excel":
        return _export_excel(export_data, title, now_str)
    elif fmt == "pdf":
        return _export_pdf(export_data, title, now_str)


def _build_export_data(stats: dict) -> dict:
    """构建导出数据字典"""
    return {
        "date": datetime.now().strftime("%Y-%m-%d"),
        "queries": stats.get("vectors", {}).get("embeddings", 0),
        "documents": stats.get("documents", {}).get("total_documents", 0),
        "users": stats.get("users", {}).get("total_users", 0),
        "storage": stats.get("vectors", {}).get("storage_size_mb", 0),
        "vectors": stats.get("vectors", {}).get("vectors_total", 0),
        "active_chunks": stats.get("documents", {}).get("active_chunks", 0),
        "collections": stats.get("vectors", {}).get("collections", 0),
        "upload_files": stats.get("storage", {}).get("total_files", 0),
        "audit_events": stats.get("audit", {}).get("total_events", 0),
    }


def _export_json(data: dict, title: str, now_str: str):
    """导出 JSON 格式"""
    from fastapi.responses import Response
    export_obj = {
        "title": title,
        "generated_at": datetime.now().isoformat(),
        "data": data,
    }
    json_str = json.dumps(export_obj, ensure_ascii=False, indent=2)
    return Response(
        content=json_str.encode("utf-8"),
        media_type="application/json; charset=utf-8",
        headers={
            "Content-Disposition": f'attachment; filename="{title}.json"',
        },
    )


def _export_csv(data: dict, title: str, now_str: str):
    """导出 CSV 格式"""
    from fastapi.responses import Response
    output = io.StringIO()
    writer = csv.writer(output)
    # BOM for Excel
    output.write("\ufeff")
    writer.writerow(["指标", "数值"])
    field_labels = {
        "date": "日期", "queries": "查询量", "documents": "文档数",
        "users": "用户数", "storage": "存储(MB)", "vectors": "向量数",
        "active_chunks": "活跃分块", "collections": "集合数",
        "upload_files": "上传文件数", "audit_events": "审计事件数",
    }
    for key, val in data.items():
        label = field_labels.get(key, key)
        writer.writerow([label, val])

    csv_content = output.getvalue()
    return Response(
        content=csv_content.encode("utf-8-sig"),
        media_type="text/csv; charset=utf-8",
        headers={
            "Content-Disposition": f'attachment; filename="{title}.csv"',
        },
    )


def _export_excel(data: dict, title: str, now_str: str):
    """导出 Excel (.xlsx) 格式 — 使用 openpyxl，不可用时降级为 CSV"""
    from fastapi.responses import Response
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = title[:31]

        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, size=12, color="FFFFFF")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )

        ws.merge_cells("A1:B1")
        ws["A1"] = title
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = Alignment(horizontal="center")

        ws.merge_cells("A2:B2")
        ws["A2"] = f'生成时间: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
        ws["A2"].alignment = Alignment(horizontal="center")

        ws["A4"] = "指标"
        ws["B4"] = "数值"
        for col in ["A", "B"]:
            cell = ws[f"{col}4"]
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

        field_labels = {
            "date": "日期", "queries": "查询量", "documents": "文档数",
            "users": "用户数", "storage": "存储(MB)", "vectors": "向量数",
            "active_chunks": "活跃分块", "collections": "集合数",
            "upload_files": "上传文件数", "audit_events": "审计事件数",
        }
        row = 5
        for key, val in data.items():
            label = field_labels.get(key, key)
            ws[f"A{row}"] = label
            ws[f"B{row}"] = val
            ws[f"A{row}"].border = thin_border
            ws[f"B{row}"].border = thin_border
            row += 1

        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 15

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return Response(
            content=output.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f'attachment; filename="{title}.xlsx"',
            },
        )
    except ImportError:
        logger.warning("openpyxl 未安装，降级为 CSV 导出")
        return _export_csv(data, title, now_str)


def _export_pdf(data: dict, title: str, now_str: str):
    """导出 PDF 格式 — 使用 reportlab，不可用时降级为 JSON"""
    from fastapi.responses import Response
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.platypus import (
            SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
        )
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4, title=title)
        story = []
        styles = getSampleStyleSheet()

        _chinese_font_available = False
        try:
            for font_path in [
                r"C:\Windows\Fonts\msyh.ttc",
                r"C:\Windows\Fonts\simsun.ttc",
                r"C:\Windows\Fonts\simhei.ttf",
            ]:
                if Path(font_path).exists():
                    pdfmetrics.registerFont(TTFont("ChineseFont", font_path))
                    _chinese_font_available = True
                    break
        except Exception:
            pass

        font_name = "ChineseFont" if _chinese_font_available else "Helvetica"

        title_style = ParagraphStyle("Title_CN", parent=styles["Title"], fontName=font_name)
        story.append(Paragraph(title, title_style))
        story.append(Spacer(1, 12))
        story.append(Paragraph(
            f'生成时间: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}', styles["Normal"]
        ))
        story.append(Spacer(1, 20))

        field_labels = {
            "date": "日期", "queries": "查询量", "documents": "文档数",
            "users": "用户数", "storage": "存储(MB)", "vectors": "向量数",
            "active_chunks": "活跃分块", "collections": "集合数",
            "upload_files": "上传文件数", "audit_events": "审计事件数",
        }
        table_data = [["指标", "数值"]]
        for key, val in data.items():
            label = field_labels.get(key, key)
            table_data.append([label, str(val)])

        t = Table(table_data, colWidths=[200, 100])
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, -1), font_name),
            ("FONTSIZE", (0, 0), (-1, 0), 12),
            ("FONTSIZE", (0, 1), (-1, -1), 10),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F2F2")]),
        ]))
        story.append(t)
        doc.build(story)
        buf.seek(0)

        return Response(
            content=buf.getvalue(),
            media_type="application/pdf",
            headers={
                "Content-Disposition": f'attachment; filename="{title}.pdf"',
            },
        )
    except ImportError:
        logger.warning("reportlab 未安装，降级为 JSON 导出")
        return _export_json(data, title, now_str)


# ── 模板管理端点 ──

@router.get("/templates")
def get_templates():
    """获取报表模板列表"""
    TEMPLATES_DIR.mkdir(parents=True, exist_ok=True)
    templates = []
    try:
        for f in sorted(TEMPLATES_DIR.glob("*.json")):
            try:
                data = json.loads(f.read_text(encoding="utf-8"))
                templates.append(data)
            except json.JSONDecodeError:
                logger.warning(f"模板文件解析失败: {f}")
    except Exception as e:
        logger.warning(f"读取模板列表失败: {e}")
    return templates


@router.post("/templates")
def create_template(request: TemplateCreateRequest):
    """创建报表模板"""
    TEMPLATES_DIR.mkdir(parents=True, exist_ok=True)
    template_id = uuid.uuid4().hex[:12]
    now = datetime.now().isoformat()
    template = {
        "id": template_id,
        "name": request.name,
        "description": request.description,
        "default_fields": request.default_fields,
        "default_format": request.default_format,
        "created_at": now,
        "updated_at": now,
    }
    file_path = TEMPLATES_DIR / f"{template_id}.json"
    file_path.write_text(json.dumps(template, ensure_ascii=False, indent=2), encoding="utf-8")
    logger.info(f"模板已创建: {template_id}")
    return template


@router.put("/templates/{template_id}")
def update_template(template_id: str, request: TemplateUpdateRequest):
    """更新报表模板"""
    file_path = TEMPLATES_DIR / f"{template_id}.json"
    if not file_path.exists():
        raise HTTPException(404, "模板不存在")

    template = json.loads(file_path.read_text(encoding="utf-8"))
    if request.name is not None:
        template["name"] = request.name
    if request.description is not None:
        template["description"] = request.description
    if request.default_fields is not None:
        template["default_fields"] = request.default_fields
    if request.default_format is not None:
        template["default_format"] = request.default_format
    template["updated_at"] = datetime.now().isoformat()

    file_path.write_text(json.dumps(template, ensure_ascii=False, indent=2), encoding="utf-8")
    return template


@router.delete("/templates/{template_id}")
def delete_template(template_id: str):
    """删除报表模板"""
    file_path = TEMPLATES_DIR / f"{template_id}.json"
    if not file_path.exists():
        raise HTTPException(404, "模板不存在")
    file_path.unlink()
    return {"status": "deleted", "id": template_id}


# ── 报表分享端点 ──

@router.post("/reports/share")
def share_report(request: ShareRequest):
    """生成报表分享链接，返回分享 URL 和 token"""
    if not request.permissions:
        raise HTTPException(400, "请至少选择一种分享权限")

    for p in request.permissions:
        if p not in ("view", "edit", "download"):
            raise HTTPException(400, f"无效权限: {p}，支持 view/edit/download")

    token = uuid.uuid4().hex
    now = datetime.now()

    expires_at = request.expires_at
    if not expires_at:
        expires_at = (now + timedelta(days=7)).isoformat()
    else:
        try:
            expires_dt = datetime.fromisoformat(expires_at)
            if expires_dt <= now:
                raise HTTPException(400, "过期时间必须在当前时间之后")
            max_expiry = now + timedelta(days=30)
            if expires_dt > max_expiry:
                raise HTTPException(400, "分享有效期不能超过 30 天")
        except ValueError:
            raise HTTPException(400, "过期时间格式无效")

    SHARED_REPORTS_DIR.mkdir(parents=True, exist_ok=True)

    share_record = {
        "token": token,
        "report_id": request.report_id,
        "permissions": request.permissions,
        "expires_at": expires_at if isinstance(expires_at, str) else expires_at.isoformat(),
        "password": request.password,
        "note": request.note,
        "created_at": now.isoformat(),
    }

    file_path = SHARED_REPORTS_DIR / f"{token}.json"
    file_path.write_text(json.dumps(share_record, ensure_ascii=False, indent=2), encoding="utf-8")

    share_url = f"/shared-report/{token}"
    logger.info(f"分享链接已生成: token={token}, report_id={request.report_id}")

    return {
        "share_url": share_url,
        "token": token,
        "expires_at": share_record["expires_at"],
        "permissions": request.permissions,
        "created_at": share_record["created_at"],
    }


@router.get("/reports/{token}")
def get_shared_report(token: str, password: Optional[str] = Query(None)):
    """通过 token 访问分享的报表，校验过期时间和密码"""
    file_path = SHARED_REPORTS_DIR / f"{token}.json"
    if not file_path.exists():
        raise HTTPException(404, "分享链接不存在或已失效")

    try:
        share_record = json.loads(file_path.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        raise HTTPException(500, "分享数据损坏")

    expires_at = datetime.fromisoformat(share_record["expires_at"])
    if datetime.now() > expires_at:
        raise HTTPException(410, "分享链接已过期")

    stored_password = share_record.get("password")
    if stored_password:
        if not password:
            return {
                "password_protected": True,
                "message": "此报表需要密码访问",
                "permissions": share_record["permissions"],
            }
        if password != stored_password:
            raise HTTPException(403, "密码错误")

    report_id = share_record["report_id"]
    try:
        stats = _build_stats()
        return {
            "report_id": report_id,
            "title": f"分享报表 - {report_id}",
            "type": "shared",
            "generated_at": share_record["created_at"],
            "sections": [{
                "title": "综合统计",
                "content": f"文档数: {stats.get('documents', {}).get('total_documents', 0)}，用户数: {stats.get('users', {}).get('total_users', 0)}",
                "metrics": {
                    "avg_value": stats.get("documents", {}).get("total_documents", 0),
                    "peak_value": stats.get("vectors", {}).get("embeddings", 0),
                },
            }],
            "permissions": share_record["permissions"],
            "owner_name": "系统用户",
            "password_protected": bool(stored_password),
        }
    except Exception as e:
        logger.error(f"获取分享报表失败: {e}")
        raise HTTPException(500, "获取报表数据失败")


@router.delete("/reports/share/{token}")
def revoke_share(token: str):
    """撤销分享链接"""
    file_path = SHARED_REPORTS_DIR / f"{token}.json"
    if not file_path.exists():
        raise HTTPException(404, "分享链接不存在")
    file_path.unlink()
    logger.info(f"分享已撤销: {token}")
    return {"status": "revoked", "token": token}


# ── POST /api/analytics/storage 也支持 GET（方便调试）──
@router.get("/storage")
async def get_storage_distribution_get():
    """存储分布 (GET)"""
    return await get_storage_distribution()


@router.get("/report")
# FAKE-ASYNC: 本函数标记 async 仅为接口统一，内部同步执行
async def generate_report_get(
    report_type: str = Query("summary", description="summary / detailed"),
    period: str = Query("month", description="today / week / month / all"),
):
    """统计报表 (GET)"""
    return await generate_report(
        ReportRequest(report_type=report_type, period=period)
    )


@router.get("/trends")
# FAKE-ASYNC: 本函数标记 async 仅为接口统一，内部同步执行
async def get_trends_get(
    period: str = Query("day", description="day / week / month"),
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
):
    """趋势数据 (GET)"""
    return await get_trends(
        TrendsRequest(period=period, start_date=start_date, end_date=end_date)
    )


@router.get("/export")
# FAKE-ASYNC: 本函数标记 async 仅为接口统一，内部同步执行
async def export_report_get(
    format: str = Query("csv"),
):
    """导出报表 (GET) — 向后兼容"""
    return await export_report_v2(
        ExportRequest(format=format, fields=[])
    )
